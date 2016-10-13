# INFORMATION: This script is designed to read the Sydian template, parse through each row, and then send an e-mail to
# the relevant users identified in that row

import openpyxl  # To modify excel
import win32com.client as win32  # To work with outlook
import sys  # For system exit functionality
import re

# Eliminates duplicates in each row of the string


def uniquify(string):
    output = ""
    seen = set()
    for row in string.split("\n"):
        seen = set()
        for word in row.split():
            if word not in seen:
                output += (word + " ")
                seen.add(word)
        output += "\n"
    return output

# ask for name of the excel file
while True:
    print("What is the name of the excel file that you are using (format: example.xlsx)?")
    name = input()

    # Load workbook with given name (must be in same folder) and worksheet
    try:
        wb = openpyxl.load_workbook(name)
        ws = wb.active
        break
    except:  # errors out if wrong file format or name
        print("Could not load your excel workbook.")


# Gather data about the Sydian spreadsheet (Note: SubHeaders may be
# removed for a generic function)
print("Which row is the header information (Division, phone#, user, employee id, etc.) in? The default value is 4.")

try:
    HeaderRow = int(input())
except:
    print("Blank or invalid entry, the program will proceed with using 4.\n")
    HeaderRow = 4

SubHeaderRow = HeaderRow + 1  # subheadings
SubSubHeaderRow = HeaderRow + 2  # sub sub headings

# Specify the first row of data to parse through
print("What is the first row in which there is user data? The default value is 7.")

try:
    FirstRow = int(input())
except:
    print("Blank or invalid entry, the program will proceed with using 7.\n")
    FirstRow = 7

# Specificy which column contains the users unique identified to convert
# into an email address
print("Which column(enter an integer value ex. A is 1, B is 2, etc.) is the SSO in? The default value is 3.")
try:
    SSO = int(input())
except:
    print("Blank or invalid entry, the program will proceed with using 3.\n")
    SSO = 3

# value to append is set here
print("This program will append '@ge.com' to the SSO column and send an e-mail to that address.")
emailappend = "@ge.com"

# Calcualte the max row and column used in the spreadsheet
MaxRow = ws.max_row
MaxColumn = ws.max_column
print("We have detected " + str(MaxRow) + " rows and " +
      str(MaxColumn) + " columns in total.")

# Starting Sequence
print("Press enter to view a sample message.")
start = input()

# Definitions
headercount = 0  # to deals with cells where the header is merged; keeps track of which merge entry we're on
count = 0  # to deal with merged subheaders
UserUsage = ""  # Create an empty string to save the message body

# iterate across all the rows starting at specified start row
for rows in range(FirstRow, MaxRow + 1, 1):

    # If there is no e-mail address/SSO to use, skip the iteration
    if (str(ws.cell(row=rows, column=SSO).value) == "None"):
        continue

    # appends the e-mail address
    recipient = str(ws.cell(row=rows, column=SSO).value) + emailappend

    # iterates through every column
    for columns in range(1, MaxColumn + 1, 1):

        # Variables to simplify visibility for code below
        data = str(ws.cell(row=rows, column=columns).value)
        headerstring = str(ws.cell(row=HeaderRow, column=columns).value)

        # NOTE: The code below is specifically designed for the Sydian template, a general case would not need any
        # subheaders or subsubheaders or workarounds for merging. Simply form the UserUsage string as follows:
        # UserUsage+=(headerstring + ": " + data + "\n")

        # Special considerations because of subheaders and subsubheaders
        subheaderstring = str(ws.cell(row=SubHeaderRow, column=columns).value)
        subsubheaderstring = str(
            ws.cell(row=SubSubHeaderRow, column=columns).value)

        # skip adding column data if the column entry is blank
        if (data == "None"):
            headercount += 1  # tracks changes to merged headers/subheaders
            count += 1
            continue

        # case where there are no subheaders or subsubheaders
        elif (subheaderstring == "None" and subsubheaderstring == "None"):
            UserUsage += (headerstring + ": " + data + "\n")

        # case where both the header and subheader1 are merged but it's not the
        # first entry
        elif (subheaderstring == "None" and headerstring == "None"):
            UserUsage += (str(ws.cell(row=HeaderRow, column=columns - headercount).value) + " " + str(
                ws.cell(row=SubHeaderRow,
                        column=columns - count).value) + " " + subsubheaderstring + ": " + data + "\n")
            count += 1
            headercount += 1

        # case where the only header row is merged and it's not the first entry
        elif (headerstring == "None"):
            UserUsage += (str(ws.cell(row=HeaderRow,
                                      column=columns - headercount).value) + " " + subheaderstring + " " + subsubheaderstring + ": " + data + "\n")
            count = 1
            headercount += 1

        # general case where the all headerrow, subheader, and subsubheader are
        # defined
        else:
            UserUsage += (headerstring + " " + subheaderstring +
                          " " + subsubheaderstring + ": " + data + "\n")
            count = 1
            headercount = 1

    # Ignore a blank string / do not attempt to build a message
    if (UserUsage == ""):
        continue

    # Create the outlook message
    outlook = win32.Dispatch("outlook.application")
    mail = outlook.CreateItem(0)
    mail.To = recipient
    mail.Subject = "Wireless Overusage Notification"

    # clean up Sydian formatting
    UserUsageSend = str(UserUsage).replace('None', '')
    UserUsageSend = UserUsageSend.replace("  ", " ")
    UserUsageSend = UserUsageSend.replace(' :', ':')
    UserUsageSend = UserUsageSend.replace(
        'Data Total Charge', '\nData Total Charge')
    UserUsageSend = UserUsageSend.replace('Taxes GST:', '\nTaxes GST:')
    UserUsageSend = UserUsageSend.replace('Charge: ', 'Charge: $')
    UserUsageSend = UserUsageSend.replace('Savings: ', 'Savings: $')
    UserUsageSend = UserUsageSend.replace('Total: ', 'Total: $')
    UserUsageSend = UserUsageSend.replace('ST: ', 'ST: $')
    UserUsageSend = uniquify(UserUsageSend)

    regex = re.compile(r'User: (\w*)')
    NameFind = regex.search(UserUsageSend)
    Name = NameFind.group(1)
    NameFix = Name.title()
    regex = re.compile(r'(Vendor: \w* ?(\w*)?)')
    vendorFind = regex.search(UserUsageSend)
    vendorFix = vendorFind.group(1)

    UserFinal = 'Hello ' + NameFix + \
        ',\n\nYour device is listed among the highest data usage at GE Canada. This notification is a reminder that data is intended for business use and should not be overused for personal matters.\n\n---------------------------\n'

    # Separates giant string by line berak, takes only the top details of the user
    # Then adds in the Roaming Costs/Usage (ignoring savings)
    MessageList = UserUsageSend.split(sep='\n')
    MessageRoaming = [
        row for row in MessageList if 'Charge' not in row if 'Vendor' not in row]
    MessageString = UserFinal + vendorFix + '\n' + '\n'.join(MessageRoaming)

    # adding an extra line break for readibility
    MessageString += '\n\n-Chris Dion (Client Services Leader)\n-Vivian Xiao (Sourcing Leader)'
    mail.body = str(MessageString)

    # Verify with the user whether they truly want to send out the e-mails
    # based on a printed sample
    if (rows == FirstRow):  # Samples only the first row of data
        print("\n" + str(MessageString))
        print("\nShown above is the sample message. Are you sure you want to send?"
              " (reply 'yes' or this script will quit)? This prompt message will not"
              " be repeated for the rest of the messages.")
        emailmode = input()
        # Exit if user types anything but yes case-insensitive
        if (emailmode.upper() != "YES"):
            sys.exit(130)

    # Send mail
    mail.send

    # Reset user message and move on to the next row
    UserUsage = ""

# On script completion
print("Complete. Enter anything to exit.")
start = input()
sys.exit(130)
