#INFORMATION: This script is designed to read the Sydian template, parse through each row, and then send an e-mail to
#the relevant users identified in that row

#Eliminates duplicates in each row of the string
def uniquify(string):
    output = ""
    seen = set()
    for row in string.split("\n"):
        seen = set()
        for word in row.split():
            if word not in seen:
                output += (word + " ")
                seen.add(word)
        output+= "\n"
    return output

import openpyxl #To modify excel
import win32com.client as win32 #To work with outlook
import sys #For system exit functionality

#ask for name of the excel file
print("What is the name of the excel file that you are using (format: example.xlsx)?")
name = input()

#Load workbook with given name (must be in same folder) and worksheet
try:
    wb = openpyxl.load_workbook(name)
    ws = wb.active
except: #errors out if wrong file format or name
        print("Could not load your excel workbook. Enter any key to exit.")
        start = input()
        sys.exit(130)

#Gather data about the Sydian spreadsheet (Note: SubHeaders may be removed for a generic function)
print("Which row is the header information (Division, phone#, user, employee id, etc.) in? The default value is 4.")
HeaderRow = int(input())
SubHeaderRow = HeaderRow+1 #subheadings
SubSubHeaderRow = HeaderRow+2 #sub sub headings

#Specify the first row of data to parse through
print("What is the first row in which there is user data? The default value is 7.")
FirstRow = int(input())

#Specificy which column contains the users unique identified to convert into an email address
print("Which column(enter an integer value ex. A is 1, B is 2, etc.) is the SSO in? The default value is 4.")
SSO = int(input())

#value to append is set here
print("This program will append '@ge.com' to the SSO column and send an e-mail to that address.")
emailappend = "@ge.com"

#Calcualte the max row and column used in the spreadsheet
MaxRow = ws.max_row
MaxColumn = ws.max_column
print("We have detected " + str(MaxRow) + " rows and " + str(MaxColumn) + " columns in total.")

#Starting Sequence
print("To start, type anything and hit enter")
start = input()

#Definitions
headercount = 0 #to deals with cells where the header is merged; keeps track of which merge entry we're on
count = 0 #to deal with merged subheaders
UserUsage = "" #Create an empty string to save the message body

for rows in range (FirstRow, MaxRow+1,1): #iterate across all the rows starting at specified start row

    #If there is no e-mail address/SSO to use, skip the iteration
    if (str(ws.cell(row=rows,column=SSO).value) == "None"):
        continue

    #appends the e-mail address
    recipient = str(ws.cell(row=rows,column=SSO).value) + emailappend

    # iterates through every column
    for columns in range(1,MaxColumn+1,1):

        # Variables to simplify visibility for code below
        data = str(ws.cell(row=rows, column=columns).value)
        headerstring = str(ws.cell(row=HeaderRow, column=columns).value)

        # NOTE: The code below is specifically designed for the Sydian template, a general case would not need any
        # subheaders or subsubheaders or workarounds for merging. Simply form the UserUsage string as follows:
        # UserUsage+=(headerstring + ": " + data + "\n")

        #Special considerations because of subheaders and subsubheaders
        subheaderstring = str(ws.cell(row=SubHeaderRow,column=columns).value)
        subsubheaderstring = str(ws.cell(row=SubSubHeaderRow,column=columns).value)

        # skip adding column data if the column entry is blank
        if (data == "None"):
            headercount+=1 #tracks changes to merged headers/subheaders
            count+=1
            continue

        #case where there are no subheaders or subsubheaders
        elif (subheaderstring == "None" and subsubheaderstring == "None" ):
            UserUsage+=(headerstring + ": " + data + "\n")

        #case where both the header and subheader1 are merged but it's not the first entry
        elif (subheaderstring == "None" and headerstring == "None"):
            if subsubheaderstring == "None": #special case within this case
                subsubheaderstring = ""
            UserUsage+=(str(ws.cell(row=HeaderRow, column=columns-headercount).value) + " " + str(
                ws.cell(row=SubHeaderRow, column=columns-count).value) + " " + subsubheaderstring+ ": " + data + "\n")
            count+=1
            headercount+=1

        #case where the only header row is merged and it's not the first entry
        elif (headerstring == "None"):
            if (subheaderstring == "None"): #special cases within this case
                subheaderstring = ""
            if subsubheaderstring == "None": #special cases within this case
                subsubheaderstring = ""
            UserUsage+=(str(ws.cell(row=HeaderRow, column=columns - headercount).value) + " " + subheaderstring + " " + subsubheaderstring + ": " + data + "\n")
            count = 1
            headercount+=1

        #general case where the all headerrow, subheader, and subsubheader are defined
        else:
            if subheaderstring == "None": #special cases within this case
                subheaderstring = ""
            if subsubheaderstring == "None": #special cases within this case
                subsubheaderstring = ""
            UserUsage+=(headerstring + " " + subheaderstring + " " + subsubheaderstring + ": " + data + "\n")
            count = 1
            headercount = 1

    #Ignore a blank string / do not attempt to build a message
    if (UserUsage == ""):
        continue

    #Create the outlook message
    outlook = win32.Dispatch("outlook.application")
    mail = outlook.CreateItem(0)
    mail.To = recipient
    mail.Subject = "Notification"

    #clean up Sydian formatting
    UserUsageSend = str(UserUsage).replace("  ", " ")
    UserUsageSend = UserUsageSend.replace(' :', ':')
    UserUsageSend = UserUsageSend.replace('Charge: ', 'Charge: $')
    UserUsageSend = UserUsageSend.replace('Savings: ', 'Savings: $')
    UserUsageSend = UserUsageSend.replace('Total: ', 'Total: $')
    UserUsageSend = UserUsageSend.replace('ST: ', 'ST: $')
    UserUsageSend = uniquify(UserUsageSend)

    mail.body = str(UserUsageSend)

    #Verify with the user whether they truly want to send out the e-mails based on a printed sample
    if (rows == FirstRow):
        print("The first e-mail will look like the following:\n" + str(UserUsageSend))
        print("Are you sure you want to send (type 'yes' or this script will quit)? This message will not be repeated.")
        emailmode = input()
        #Exit if user types anything but yes case-insensitive
        if (emailmode.upper() != "YES"):
            sys.exit(130)

    #Send mail
    mail.send

    #Reset user message and move on to the next row
    UserUsage=""

#On script completion
print("Complete. Enter anything to exit.")
start = input()
sys.exit(130)


