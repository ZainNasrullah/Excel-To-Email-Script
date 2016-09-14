#INFORMATION: This script is designed to read the Sydian template, parse through each row, and then send an e-mail to
#the relevant users identified in that row

#Eliminates duplicates in each row of the string
def uniquify(string):
    output = ""
    seen = set()
    for row in string.split("\n"):
        seen = set()
        for word in row.split():
            if word not in seen:
                output += (word + " ")
                seen.add(word)
        output+= "\n"
    return output

import openpyxl #To modify excel
import win32com.client as win32 #To work with outlook
import sys #For system exit functionality

#ask for name of the excel file
print("What is the name of the excel file that you are using (format: example.xlsx)?")
name = input()

#Load workbook with given name (must be in same folder) and worksheet
try:
    wb = openpyxl.load_workbook(name)
    ws = wb.active
except: #errors out if wrong file format or name
        print("Could not load your excel workbook. Enter any key to exit.")
        start = input()
        sys.exit(130)

#Gather data about the Sydian spreadsheet (Note: SubHeaders may be removed for a generic function)
print("Which row is the header information in?")
HeaderRow = int(input())

#Specify the first row of data to parse through
print("What is the first row in which there is user data?")
FirstRow = int(input())

#Specificy which column contains the users unique identified to convert into an email address
print("Which column(enter an integer value) is the SSO/e-mail address in?")
SSO = int(input())

#value to append is set here
print("Do you want to append something to the SSO column(ex. append '@gmail.com')? Yes or no?")
start = input()
if start.upper() == "YES":
    print ("What do you want to append?")
    emailappend = input()
    print ("The e-mail format will be example"+emailappend+'.')
else:
    print("Nothing will be appended.")
    emailappend = ""

#Calcualte the max row and column used in the spreadsheet
MaxRow = ws.max_row
MaxColumn = ws.max_column
print("We have detected " + str(MaxRow) + " rows and " + str(MaxColumn) + " columns in total.")

#Starting Sequence
print("To start, type anything and hit enter")
start = input()

#Definitions
UserUsage = "" #Create an empty string to save the message body

for rows in range (FirstRow, MaxRow+1,1): #iterate across all the rows starting at specified start row

    #If there is no e-mail address/SSO to use, skip the iteration
    if (str(ws.cell(row=rows,column=SSO).value) == "None"):
        continue

    #appends the e-mail address
    recipient = str(ws.cell(row=rows,column=SSO).value) + emailappend

    # iterates through every column
    for columns in range(1,MaxColumn+1,1):

        # Variables to simplify visibility for code below
        data = str(ws.cell(row=rows, column=columns).value)
        headerstring = str(ws.cell(row=HeaderRow, column=columns).value)

        # NOTE: The code below is specifically designed for the Sydian template, a general case would not need any
        # subheaders or subsubheaders or workarounds for merging. Simply form the UserUsage string as follows:
        # UserUsage+=(headerstring + ": " + data + "\n")

        # skip adding column data if the column entry is blank
        if (data == "None"):
            continue
        #Data exists
        else:
            UserUsage+=(headerstring + ": " + data + "\n")

    #Ignore a blank string / do not attempt to build a message
    if (UserUsage == ""):
        continue

    #Create the outlook message
    outlook = win32.Dispatch("outlook.application")
    mail = outlook.CreateItem(0)
    mail.To = recipient
    mail.Subject = "Notification"

    #clean up formatting
    UserUsageSend = str(UserUsage).replace("  ", " ")

    #Below can remove non-unique entries within each row if it looks unsightly
    #UserUsageSend = uniquify(UserUsageSend)

    mail.body = str(UserUsageSend)

    #Verify with the user whether they truly want to send out the e-mails based on a printed sample
    if (rows == FirstRow):
        print("\n" + str(UserUsageSend))
        print("Shown above is a sample message. Are you sure you want to send?"
              " (reply 'yes' or this script will quit)? This prompt message will not"
              " be repeated for the rest of the messages.")
        emailmode = input()
        #Exit if user types anything but yes case-insensitive
        if (emailmode.upper() != "YES"):
            sys.exit(130)

    #Send mail
    mail.send

    #Reset user message and move on to the next row
    UserUsage=""

#On script completion
print("Complete. Enter anything to exit.")
start = input()
sys.exit(130)


